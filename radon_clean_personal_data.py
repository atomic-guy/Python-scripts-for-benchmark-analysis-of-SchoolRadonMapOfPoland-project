import os
from radon_statistics import getListofExcelFiles
from openpyxl import load_workbook
import win32com.client as win32

def xls_to_xlsx(ListofExcelFiles):
    '''
    Saves Excel files with extension .xls as .xlsx.  
            Parameters:
                    ListofExcelFiles (list of str): List of .xls files with full path
    '''
    for file in ListofExcelFiles:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(file)

        wb.SaveAs(file+"x", FileFormat = 51)     #FileFormat = 51 is for .xlsx extension
        wb.Close()                               #FileFormat = 56 is for .xls extension
        excel.Application.Quit()
    print("All files from the list were saved as .xlsx files.")
    return

def xlsx_to_xls(ListofExcelFiles):
    '''
    Saves Excel files with extension .xlsx as .xls.  
            Parameters:
                    ListofExcelFiles (list of str): List of .xlsx files with full path
    '''
    for file in ListofExcelFiles:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(file)
        new_file = file[:-1]
        wb.SaveAs(new_file, FileFormat = 56)    #FileFormat = 51 is for .xlsx extension
        wb.Close()                              #FileFormat = 56 is for .xls extension
        excel.Application.Quit()
    print("All files from the list were saved as .xls files.")
    return

def removeFiles(ListofFiles):
    '''
    Removes all files from the list. 
            Parameters:
                    ListofFiles (list of str): List of files with full path (if python script 
                                        is running from other directory then files) to remove
    '''
    for file in ListofFiles:
        os.remove(file)
    print("All files from the list were removed.")
    return

def cleanPersonalDataInExcelFiles(ListofExcelFiles):
    '''
    Cleans personal data in Excel files. All files must have .xlsx extension.
            Parameters:
                    ListofExcelFiles (list of str): List of .xlsx files with full path
    '''
    for file in ListofExcelFiles:
        workbook = load_workbook(file)
        sheet = workbook['Informacje ogólne']
        
        file_name = os.path.basename(file)
        id = int(file_name[0:2])
        #School
        school = "Szkoła Nr " + str(id)
        sheet['C3'] = school       
        #City
        sheet['C4'] = "Miejscowość"
        #Teacher
        sheet['C6'] = "Nauczyciel"
        #Students
        cell_list = list(range(8,30))
        n = 1
        for i in cell_list:
            if sheet.cell(column=3, row=i).value:
                uczen = "Uczeń " + str(n)
                sheet.cell(column=3, row=i).value = uczen
                n = n + 1
        workbook.save(file)
    print("The task is done")
    return


DirectoryPath = "E:\\--PYTHON\\Skrypty\\Radon\\results_xls_files\\"
ListofExcelFiles = getListofExcelFiles(DirectoryPath, 'xlsx')

#removeFiles(ListofExcelFiles)
#xls_to_xlsx(ListofExcelFiles)
#cleanPersonalDataInExcelFiles(ListofExcelFiles)
#xlsx_to_xls(ListofExcelFiles)